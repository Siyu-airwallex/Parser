from openpyxl import load_workbook
import sys

reload(sys)
sys.setdefaultencoding('utf-8')
wb = load_workbook('Source/Template3.xlsx')

print wb.get_sheet_names()

matrix = wb['stefanie']


def getRequiredFeilds(payerCountry, beneficiaryCountry, transactionType, channel, paymentMethod, paymentCountry,
                      paymentCurrency):
  lst = []
  for col in matrix.iter_cols(min_row=1, max_row=7, min_col=3):
    if(col[0].value == None): break
    payerCountryList = col[0].value.split(",")
    beneficiaryCountryList = col[1].value.split(",")
    transactionTypeList = col[2].value.split(",")
    channelList = col[3].value.split(",")
    paymentMethodList = col[4].value.split(",")
    paymentCountryList = col[5].value.split(",")
    paymentCurrencyList = col[6].value.split(",")
    if( (payerCountry == '*' or payerCountry in payerCountryList or "*" in payerCountryList) and
        (beneficiaryCountry == '*' or beneficiaryCountry in beneficiaryCountryList or "*" in beneficiaryCountryList) and
        (transactionType == '*' or transactionType in transactionTypeList or "*" in transactionTypeList) and
        (channel == '*' or channel in channelList or "*" in channelList) and
        (paymentMethod == '*' or paymentMethod in paymentMethodList or "*" in paymentMethodList) and
        (paymentCountry == '*' or paymentCountry in paymentCountryList or "*" in paymentCountryList) and
        (paymentCurrency == '*' or paymentCurrency in paymentCurrencyList or "*" in paymentCurrencyList)
      ):
       lst.append(col[1].col_idx)

  dic = {}
  for col in lst:
    tmp = []
    for row in range(8, matrix.max_row):
      value = str(matrix.cell(row=row, column=col).value)
      if value.startswith('M'):
        tmp.append(row)
    dic[col] = tmp
  return dic




condition = "When Payer country is %s, Beneficiary country is %s, Transaction type is %s, Channel is %s, Payment method is %s, Payment country is %s, Payment currency is %s: \n"
output = "Required fields are:"

def dumpRequiredFields(dic):
  for key, value in dic.items():
    payerCountry = matrix.cell(row=1, column=key).value
    beneficiaryCountry = matrix.cell(row=2, column=key).value
    transactionType = matrix.cell(row=3, column=key).value
    channel = matrix.cell(row=4, column=key).value
    paymentMethod = matrix.cell(row=5, column=key).value
    paymentCountry = matrix.cell(row=6, column=key).value
    paymentCurrency = matrix.cell(row=7, column=key).value
    print(condition % (payerCountry, beneficiaryCountry, transactionType, channel, paymentMethod, paymentCountry, paymentCurrency))
    print output
    for row in value:
      print matrix.cell(row=row, column=2).value + "\t"
    print "\n"


payerCountry = raw_input("Please enter payer country: ")
beneficiaryCountry = raw_input("Please enter beneficiary country: ")
transactionType = raw_input("Please enter transaction type: ")
channel = raw_input("Please enter channel: ")
paymentMethod = raw_input("Please enter payment method: ")
paymentCountry = raw_input("Please enter payment country: ")
paymentCurrency = raw_input("Please enter payment currency: ")

# print transactionType, paymentMethod, paymentCountry, paymentCurrency

res = getRequiredFeilds(payerCountry, beneficiaryCountry, transactionType, channel, paymentMethod, paymentCountry, paymentCurrency)

dumpRequiredFields(res)
