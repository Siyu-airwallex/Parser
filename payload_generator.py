import json
from dotmap import DotMap
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook
from datetime import datetime
from datetime import timedelta

import random
import uuid
import sys

reload(sys)
sys.setdefaultencoding('utf-8')
wb = load_workbook('Source/Template3.xlsx')

print wb.get_sheet_names()

matrix = wb['stefanie']
print (matrix['A1'].value)

payerCountry = raw_input("Please enter payer country: ")
beneficiaryCountry = raw_input("Please enter beneficiary country: ")
transactionType = raw_input("Please enter transaction type: ")
channel = raw_input("Please enter channel: ")
paymentMethod = raw_input("Please enter payment method: ")
paymentCountry = raw_input("Please enter payment country: ")
paymentCurrency = raw_input("Please enter payment currency: ")

def getRequiredFeilds():
  lst = []
  for col in matrix.iter_cols(min_row=1, max_row=7, min_col=3):
    if(col[0].value == None): break
    payerCountryList = col[0].value.split(",")
    beneficiaryCountryList = col[1].value.split(",")
    transactionTypeList = col[2].value.split(",")
    channelList = col[3].value.split(",")
    paymentMethodList = col[4].value.split(",")
    paymentCountryList = col[5].value.split(",")
    paymentCurrencyList = col[6].value.split(",")
    if( (payerCountry == '*' or payerCountry in payerCountryList or "*" in payerCountryList) and
        (beneficiaryCountry == '*' or beneficiaryCountry in beneficiaryCountryList or "*" in beneficiaryCountryList) and
        (transactionType == '*' or transactionType in transactionTypeList or "*" in transactionTypeList) and
        (channel == '*' or channel in channelList or "*" in channelList) and
        (paymentMethod == '*' or paymentMethod in paymentMethodList or "*" in paymentMethodList) and
        (paymentCountry == '*' or paymentCountry in paymentCountryList or "*" in paymentCountryList) and
        (paymentCurrency == '*' or paymentCurrency in paymentCurrencyList or "*" in paymentCurrencyList)
      ):
       lst.append(col[1].col_idx)

  dic = {}
  for col in lst:
    tmp = []
    for row in range(8, matrix.max_row):
      value = str(matrix.cell(row=row, column=col).value)
      if value.startswith('M'):
        tmp.append(row)
    dic[col] = tmp
  return dic


def generateTest(dic):
    for key,value in dic.items():
        tmp = DotMap()
        baseFieldsValuation(key,tmp,dicName="tmp")
        otherFieldsValuation(value,tmp,dicName="tmp")
        json_data = json.dumps(tmp.toDict())
        print json_data

BASE_FIELDS_LIST = []


def baseFieldsValuation(key,tmp,dicName):
    global payerCountry, beneficiaryCountry, transactionType, channel, paymentMethod, paymentCountry, paymentCurrency
    (payerCountryList, beneficiaryCountryList, transactionTypeList, channelList, paymentMethodList, paymentCountryList, paymentCurrencyList) = (matrix.cell(column=key, row=i).value.split(',') for i in range(1,8))

    if payerCountry != '*':
      exec (dicName + ".payer.address.country_code = payerCountry") in globals(), locals()
      BASE_FIELDS_LIST.append("payer.address.country_code")
    elif '*' not in payerCountryList:
      payerCountry = random.choice(payerCountryList)
      exec (dicName + ".payer.address.country_code = payerCountry") in globals(), locals()
      BASE_FIELDS_LIST.append("payer.address.country_code")

    if beneficiaryCountry != "*":
      exec(dicName + ".beneficiary.address.country_code = beneficiaryCountry") in globals(), locals()
      BASE_FIELDS_LIST.append("beneficiary.address.country_code")
    elif '*' not in beneficiaryCountryList:
      beneficiaryCountry = random.choice(beneficiaryCountryList)
      exec(dicName + ".beneficiary.address.country_code = beneficiaryCountry") in globals(), locals()
      BASE_FIELDS_LIST.append("beneficiary.address.country_code")

    if transactionType != '*':
      transTypeList = transactionType.split('2')
      exec(dicName + ".payer.entity_type = " + " 'PERSONAL' if transTypeList[0] == 'P' else 'COMPANY' ") in globals(),locals()
      exec(dicName + ".beneficiary.entity_type = " + " 'PERSONAL' if transTypeList[1] == 'P' else 'COMPANY' ") in globals(),locals()
      BASE_FIELDS_LIST.append("beneficiary.entity_type")
      BASE_FIELDS_LIST.append("payer.entity_type")
    elif '*' not in transactionTypeList:
      transactionType = random.choice(transactionTypeList)
      transTypeList = transactionType.split('2')
      exec(dicName + ".payer.entity_type = " + " 'PERSONAL' if transTypeList[0] == 'P' else 'COMPANY' ") in globals(),locals()
      exec(dicName + ".beneficiary.entity_type = " + " 'PERSONAL' if transTypeList[1] == 'P' else 'COMPANY' ") in globals(),locals()
      BASE_FIELDS_LIST.append("beneficiary.entity_type")
      BASE_FIELDS_LIST.append("payer.entity_type")

    if paymentMethod != '*':
      exec(dicName + ".payment_method = paymentMethod" ) in globals(),locals()
      BASE_FIELDS_LIST.append("payment_method")
    elif '*' not in paymentMethodList:
      paymentMethod = random.choice(paymentMethodList)
      exec(dicName + ".payment_method = paymentMethod" ) in globals(),locals()
      BASE_FIELDS_LIST.append("payment_method")

    if paymentCountry != '*':
      exec(dicName + ".beneficiary.bank_details.bank_country_code = paymentCountry" ) in globals(),locals()
      BASE_FIELDS_LIST.append("beneficiary.bank_details.bank_country_code")
    elif '*' not in paymentCountryList:
      paymentCountry = random.choice(paymentCountryList)
      exec(dicName + ".beneficiary.bank_details.bank_country_code = paymentCountry" ) in globals(),locals()
      BASE_FIELDS_LIST.append("beneficiary.bank_details.bank_country_code")

    if paymentCurrency != '*':
      exec (dicName + ".payment_currency = paymentCurrency") in globals(), locals()
      exec (dicName + ".beneficiary.bank_details.account_currency = paymentCurrency") in globals(), locals()
      BASE_FIELDS_LIST.append("payment_currency")
      BASE_FIELDS_LIST.append("beneficiary.bank_details.account_currency")
    elif '*' not in paymentCountryList:
      paymentCountry = random.choice(paymentCountryList)
      exec (dicName + ".payment_currency = paymentCurrency") in globals(), locals()
      exec (dicName + ".beneficiary.bank_details.account_currency = paymentCurrency") in globals(), locals()
      BASE_FIELDS_LIST.append("payment_currency")
      BASE_FIELDS_LIST.append("beneficiary.bank_details.account_currency")

    exec(dicName + ".request_id = str(uuid.uuid4())") in globals(), locals()
    BASE_FIELDS_LIST.append("request.id")
    exec(dicName + ".payment_date = str((datetime.now() + timedelta(days=2)).date())") in globals(), locals()
    BASE_FIELDS_LIST.append("payment_date")



def otherFieldsValuation(val,tmp,dicName):
    fieldsInLayerName = [matrix.cell(column=2, row=i).value for i in val]
    #print len(fieldsInLayerName)
    with open("test_datastore/payload.json") as data_file:
        data = json.load(data_file)
    for fieldName in fieldsInLayerName:
        if fieldName not in BASE_FIELDS_LIST:
            fieldValue = getValueFromDataset(fieldName, data)
            exec(dicName + "." + fieldName + " =  fieldValue") in globals(), locals()
    data_file.close()


def getValueFromDataset(fieldName, jsondata):
    layers = fieldName.split('.')
    curr = jsondata
    for layer in layers:
        curr = curr[layer]
    return random.choice(curr)



res = getRequiredFeilds()


generateTest(res)