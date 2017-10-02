from openpyxl import load_workbook
import sys
import json
import random
from dotmap import DotMap
from Model.Transaction import Transaction
import uuid
from datetime import datetime
from datetime import timedelta

class Parser:
  matrix = None
  transaction = Transaction()

  BASE_FIELDS_LIST = []

  def __init__(self, transaction):
    reload(sys)
    sys.setdefaultencoding('utf-8')
    self.matrix = load_workbook('Source/Template.xlsx')['stefanie']
    self.transaction = transaction

  def setTransaction(self, transaction):
    self.transaction = transaction

  def getRequiredFeilds(self):
    payerCountry = self.transaction.payerCountry
    beneficiaryCountry = self.transaction.beneficiaryCountry
    transactionType = self.transaction.transactionType
    channel = self.transaction.channel
    paymentMethod = self.transaction.paymentMethod
    paymentCountry = self.transaction.paymentCountry
    paymentCurrency = self.transaction.paymentCurrency

    lst = []
    for col in self.matrix.iter_cols(min_row=1, max_row=7, min_col=3):
      if (col[0].value == None): break
      payerCountryList = col[0].value.split(",")
      beneficiaryCountryList = col[1].value.split(",")
      transactionTypeList = col[2].value.split(",")
      channelList = col[3].value.split(",")
      paymentMethodList = col[4].value.split(",")
      paymentCountryList = col[5].value.split(",")
      paymentCurrencyList = col[6].value.split(",")
      if ((payerCountry == '*' or payerCountry in payerCountryList or "*" in payerCountryList) and
            (beneficiaryCountry == '*' or beneficiaryCountry in beneficiaryCountryList or "*" in beneficiaryCountryList) and
            (transactionType == '*' or transactionType in transactionTypeList or "*" in transactionTypeList) and
            (channel == '*' or channel in channelList or "*" in channelList) and
            (paymentMethod == '*' or paymentMethod in paymentMethodList or "*" in paymentMethodList) and
            (paymentCountry == '*' or paymentCountry in paymentCountryList or "*" in paymentCountryList) and
            (paymentCurrency == '*' or paymentCurrency in paymentCurrencyList or "*" in paymentCurrencyList)
          ):
        lst.append(col[1].col_idx)

    dic = {}
    for col in lst:
      tmp = []
      for row in range(8, self.matrix.max_row):
        value = str(self.matrix.cell(row=row, column=col).value)
        if value.startswith('M'):
          tmp.append(row)
      dic[col] = tmp
    return dic


  def dumpRequiredFields(self):
    dic = self.getRequiredFeilds()
    res = {}
    for key, value in dic.items():
      payerCountry = self.matrix.cell(row=1, column=key).value
      beneficiaryCountry = self.matrix.cell(row=2, column=key).value
      transactionType = self.matrix.cell(row=3, column=key).value
      channel = self.matrix.cell(row=4, column=key).value
      paymentMethod = self.matrix.cell(row=5, column=key).value
      paymentCountry = self.matrix.cell(row=6, column=key).value
      paymentCurrency = self.matrix.cell(row=7, column=key).value
      trans = Transaction(payerCountry,beneficiaryCountry, transactionType,channel,paymentMethod, paymentCountry, paymentCurrency)
      requiredFields = []
      for row in value:
        requiredFields.append(self.matrix.cell(row=row, column=2).value)
        res[trans] = requiredFields
    return res

  def generateTest(self):
    dic = self.getRequiredFeilds()
    res = {}
    for key, value in dic.items():
      payerCountry = self.matrix.cell(row=1, column=key).value
      beneficiaryCountry = self.matrix.cell(row=2, column=key).value
      transactionType = self.matrix.cell(row=3, column=key).value
      channel = self.matrix.cell(row=4, column=key).value
      paymentMethod = self.matrix.cell(row=5, column=key).value
      paymentCountry = self.matrix.cell(row=6, column=key).value
      paymentCurrency = self.matrix.cell(row=7, column=key).value
      trans = Transaction(payerCountry, beneficiaryCountry, transactionType, channel, paymentMethod, paymentCountry,paymentCurrency)
      tmp = DotMap()
      self.baseFieldsValuation(key, tmp, dicName="tmp")
      self.otherFieldsValuation(value, tmp, dicName="tmp")
      json_data = json.dumps(tmp.toDict())
      res[trans] = json_data
    return res


  def baseFieldsValuation(self, key, tmp, dicName):

    payerCountry = self.transaction.payerCountry
    beneficiaryCountry = self.transaction.beneficiaryCountry
    transactionType = self.transaction.transactionType
    paymentMethod = self.transaction.paymentMethod
    paymentCountry = self.transaction.paymentCountry
    paymentCurrency = self.transaction.paymentCurrency

    (payerCountryList, beneficiaryCountryList, transactionTypeList, channelList, paymentMethodList, paymentCountryList,
     paymentCurrencyList) = (self.matrix.cell(column=key, row=i).value.split(',') for i in range(1, 8))

    if payerCountry != '*':
      exec (dicName + ".payer.address.country_code = payerCountry") in globals(), locals()
      self.BASE_FIELDS_LIST.append("payer.address.country_code")
    elif '*' not in payerCountryList:
      payerCountry = random.choice(payerCountryList)
      exec (dicName + ".payer.address.country_code = payerCountry") in globals(), locals()
      self.BASE_FIELDS_LIST.append("payer.address.country_code")

    if beneficiaryCountry != "*":
      exec (dicName + ".beneficiary.address.country_code = beneficiaryCountry") in globals(), locals()
      self.BASE_FIELDS_LIST.append("beneficiary.address.country_code")
    elif '*' not in beneficiaryCountryList:
      beneficiaryCountry = random.choice(beneficiaryCountryList)
      exec (dicName + ".beneficiary.address.country_code = beneficiaryCountry") in globals(), locals()
      self.BASE_FIELDS_LIST.append("beneficiary.address.country_code")

    if transactionType != '*':
      transTypeList = transactionType.split('2')
      exec (
      dicName + ".payer.entity_type = " + " 'PERSONAL' if transTypeList[0] == 'P' else 'COMPANY' ") in globals(), locals()
      exec (
      dicName + ".beneficiary.entity_type = " + " 'PERSONAL' if transTypeList[1] == 'P' else 'COMPANY' ") in globals(), locals()
      self.BASE_FIELDS_LIST.append("beneficiary.entity_type")
      self.BASE_FIELDS_LIST.append("payer.entity_type")
    elif '*' not in transactionTypeList:
      transactionType = random.choice(transactionTypeList)
      transTypeList = transactionType.split('2')
      exec (
      dicName + ".payer.entity_type = " + " 'PERSONAL' if transTypeList[0] == 'P' else 'COMPANY' ") in globals(), locals()
      exec (
      dicName + ".beneficiary.entity_type = " + " 'PERSONAL' if transTypeList[1] == 'P' else 'COMPANY' ") in globals(), locals()
      self.BASE_FIELDS_LIST.append("beneficiary.entity_type")
      self.BASE_FIELDS_LIST.append("payer.entity_type")

    if paymentMethod != '*':
      exec (dicName + ".payment_method = paymentMethod") in globals(), locals()
      self.BASE_FIELDS_LIST.append("payment_method")
    elif '*' not in paymentMethodList:
      paymentMethod = random.choice(paymentMethodList)
      exec (dicName + ".payment_method = paymentMethod") in globals(), locals()
      self.BASE_FIELDS_LIST.append("payment_method")

    if paymentCountry != '*':
      exec (dicName + ".beneficiary.bank_details.bank_country_code = paymentCountry") in globals(), locals()
      self.BASE_FIELDS_LIST.append("beneficiary.bank_details.bank_country_code")
    elif '*' not in paymentCountryList:
      paymentCountry = random.choice(paymentCountryList)
      exec (dicName + ".beneficiary.bank_details.bank_country_code = paymentCountry") in globals(), locals()
      self.BASE_FIELDS_LIST.append("beneficiary.bank_details.bank_country_code")

    if paymentCurrency != '*':
      exec (dicName + ".payment_currency = paymentCurrency") in globals(), locals()
      exec (dicName + ".beneficiary.bank_details.account_currency = paymentCurrency") in globals(), locals()
      self.BASE_FIELDS_LIST.append("payment_currency")
      self.BASE_FIELDS_LIST.append("beneficiary.bank_details.account_currency")
    elif '*' not in paymentCountryList:
      paymentCountry = random.choice(paymentCountryList)
      exec (dicName + ".payment_currency = paymentCurrency") in globals(), locals()
      exec (dicName + ".beneficiary.bank_details.account_currency = paymentCurrency") in globals(), locals()
      self.BASE_FIELDS_LIST.append("payment_currency")
      self.BASE_FIELDS_LIST.append("beneficiary.bank_details.account_currency")

    exec (dicName + ".request_id = str(uuid.uuid4())") in globals(), locals()
    self.BASE_FIELDS_LIST.append("request.id")
    exec (dicName + ".payment_date = str((datetime.now() + timedelta(days=2)).date())") in globals(), locals()
    self.BASE_FIELDS_LIST.append("payment_date")

  def otherFieldsValuation(self, val, tmp, dicName):
    fieldsInLayerName = [self.matrix.cell(column=2, row=i).value for i in val]
    with open("test_datastore/payload.json") as data_file:
      data = json.load(data_file)
    for fieldName in fieldsInLayerName:
      if fieldName not in self.BASE_FIELDS_LIST:
        fieldValue = self.getValueFromDataset(fieldName, data)
        exec (dicName + "." + fieldName + " =  fieldValue") in globals(), locals()
    data_file.close()


  def getValueFromDataset(self, fieldName, jsondata):
    layers = fieldName.split('.')
    curr = jsondata
    for layer in layers:
      curr = curr[layer]
    return random.choice(curr)